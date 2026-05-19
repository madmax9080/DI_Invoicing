from datetime import date
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
from io import BytesIO
from typing import Optional
from sqlalchemy import extract, func
from fastapi import APIRouter, Depends, HTTPException, Query
from sqlalchemy.orm import Session, selectinload
from app.routers.dashboard import get_cache, get_fiscal_range
from ..dependencies.fbr import get_fbr_client_secure
from ..database import get_db
from ..models import Buyer, Invoice, InvoiceItem
from fastapi.responses import StreamingResponse
import openpyxl
import qrcode
from sqlalchemy.orm import Session
from pathlib import Path
from io import BytesIO
from app.services.pdf_service import generate_invoice_pdf_rl
from app.services.cache import get_cache, set_cache

router = APIRouter(prefix="/reports", tags=["Reports"])
BASE_DIR = Path(__file__).resolve().parent.parent
TEMPLATE_DIR = BASE_DIR / "templates"

@router.get("/sales-by-buyer")
def sales_by_buyer(
    fy: str | None = None,
    db: Session = Depends(get_db),
    fbr = Depends(get_fbr_client_secure)
):
    client_id = fbr.client_id
    cache_key = f"reports:{client_id}:sales_by_buyer:{fy or 'current'}"
    cached = get_cache(cache_key)
    if cached:
        return cached
    fiscal_start, fiscal_end, fiscal_label = get_fiscal_range(fy)
    subq = (
        db.query(
            Invoice.buyer_id.label("buyer_id"),
            func.sum(InvoiceItem.valueSalesExcludingST).label("total")
        )
        .join(InvoiceItem, Invoice.id == InvoiceItem.invoice_id)
        .filter(
            Invoice.client_id == client_id,
            Invoice.status == "posted",
            Invoice.invoiceDate.between(fiscal_start, fiscal_end)
        )
        .group_by(Invoice.buyer_id)
        .subquery()
    )
    results = (
        db.query(
            Buyer.name,
            subq.c.total
        )
        .join(subq, Buyer.id == subq.c.buyer_id)
        .order_by(subq.c.total.desc())
        .limit(15)
        .all()
    )
    output = {
        "fiscal_year": fiscal_label,
        "labels": [r.name or "Unknown" for r in results],
        "series": [float(r.total or 0) for r in results]
    }
    set_cache(cache_key, output)
    return output

# @router.get("/invoices")
# def reports_invoices(
#     fy: Optional[str] = Query(None),
#     page: int = Query(1, ge=1),
#     per_page: int = Query(25, ge=1, le=200),
#     status: Optional[str] = Query(None),
#     date_from: Optional[date] = Query(None),
#     date_to: Optional[date] = Query(None),
#     month: Optional[str] = Query(None),
#     db: Session = Depends(get_db),
#     fbr=Depends(get_fbr_client_secure)
# ):
#     client_id = fbr.client_id
#     today = date.today()
#     if fy:
#         start_year = int(fy.split("-")[0])
#     else:
#         start_year = today.year if today.month >= 7 else today.year - 1
#     fiscal_start = date(start_year, 7, 1)
#     fiscal_end = date(start_year + 1, 6, 30)
#     query = (
#         db.query(Invoice)
#         .options(joinedload(Invoice.items))
#         .filter(
#             Invoice.client_id == client_id,
#             Invoice.invoiceDate >= fiscal_start,
#             Invoice.invoiceDate <= fiscal_end
#         )
#         .order_by(Invoice.invoiceDate.desc())
#     )
#     if status:
#         query = query.filter(Invoice.status == status)
#     if month:
#         query = query.filter(
#             func.to_char(Invoice.invoiceDate, "YYYY-MM") == month
#         )
#     else:
#         if date_from:
#             query = query.filter(Invoice.invoiceDate >= date_from)
#         if date_to:
#             query = query.filter(Invoice.invoiceDate <= date_to)
#     total = query.count()
#     pages = (total + per_page - 1) // per_page
#     invoices = (
#         query
#         .offset((page - 1) * per_page)
#         .limit(per_page)
#         .all()
#     )
#     return {
#         "data": invoices,
#         "total": total,
#         "page": page,
#         "per_page": per_page,
#         "pages": pages
#     }

@router.get("/invoices")
def reports_invoices(
    fy: Optional[str] = Query(None),
    page: int = Query(1, ge=1),
    per_page: int = Query(25, ge=1, le=200),
    status: Optional[str] = Query(None),
    date_from: Optional[date] = Query(None),
    date_to: Optional[date] = Query(None),
    month: Optional[str] = Query(None),
    db: Session = Depends(get_db),
    fbr = Depends(get_fbr_client_secure)
):
    client_id = fbr.client_id
    fiscal_start, fiscal_end, _ = get_fiscal_range(fy)
    base_filter = [
        Invoice.client_id == client_id,
        Invoice.invoiceDate.between(fiscal_start, fiscal_end)
    ]
    if status:
        base_filter.append(Invoice.status == status)
    if month:
        year, month_num = map(int, month.split("-"))
        start = date(year, month_num, 1)
        end = date(year + (month_num // 12), (month_num % 12) + 1, 1)
        base_filter.append(Invoice.invoiceDate >= start)
        base_filter.append(Invoice.invoiceDate < end)
    else:
        if date_from:
            base_filter.append(Invoice.invoiceDate >= date_from)
        if date_to:
            base_filter.append(Invoice.invoiceDate <= date_to)
    total = db.query(func.count(Invoice.id)).filter(*base_filter).scalar()
    invoices = (
        db.query(Invoice)
        .options(selectinload(Invoice.items))
        .filter(*base_filter)
        .order_by(Invoice.invoiceDate.desc())
        .offset((page - 1) * per_page)
        .limit(per_page)
        .all()
    )
    return {
        "data": invoices,
        "total": total,
        "page": page,
        "per_page": per_page,
        "pages": (total + per_page - 1) // per_page
    }

@router.get("/export/excel")
def export_excel(
    fy: str,
    status: Optional[str] = None,
    date_from: Optional[date] = None,
    date_to: Optional[date] = None,
    month: Optional[int] = None,
    db: Session = Depends(get_db),
    fbr = Depends(get_fbr_client_secure)
):  
    query = db.query(Invoice).filter(
        Invoice.client_id == fbr.client_id
    )
    if status:
        query = query.filter(Invoice.status == status)
    if date_from:
        query = query.filter(Invoice.invoiceDate >= date_from)
    if date_to:
        query = query.filter(Invoice.invoiceDate <= date_to)
    if month:
        query = query.filter(
            extract("month", Invoice.invoiceDate) == month
        )
    if fy:
        start_year = int(fy.split("-")[0])
        query = query.filter(
            Invoice.invoiceDate >= date(start_year, 7, 1),
            Invoice.invoiceDate <= date(start_year + 1, 6, 30)
        )
    invoices = query.order_by(
        Invoice.invoiceDate.desc()
    ).all()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sales Report"
    headers = [
        "Date",
        "FBR Invoice No",
        "Inv. Ref No.(for Debit Inv)",
        "Buyer Business Name",
        "Buyer NTN/CNIC",
        "Reg. Status",
        "HS Code",
        "Product",
        "Qty",
        "UOM",
        "Sale Type",
        "Value Excl ST",
        "Rate",
        "Sales Tax",
        "STWH",
        "Further Tax",
        "Extra Tax",
        "FED Payable",
        "Discount",
        "SRO Schedule Number",
        "SRO Item Sr. No.",
        "Total"
    ]
    ws.append(headers)
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")
    row_num = 2
    for inv in invoices:
        for item in inv.items:
            ws.append([
                inv.invoiceDate.strftime("%d-%m-%Y"),
                inv.fbrInvoiceNo,
                inv.invoiceRefNo,
                inv.buyerBusinessName,
                inv.buyerNTNCNIC,
                inv.buyerRegistrationType,
                item.hsCode,
                item.productDescription,
                float(item.quantity or 0),
                item.uom,
                item.saleType,
                float(item.valueSalesExcludingST or 0),
                item.rate,
                float(item.salesTaxApplicable or 0),
                float(item.salesTaxWithheldAtSource or 0),
                float(item.furtherTax or 0),
                item.extraTax,
                float(item.fedPayable or 0),
                float(item.discount or 0),
                item.sroScheduleNo,
                item.sroItemSerialNo,
                float(item.totalValues or 0)
            ])
            row_num += 1
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value:
                    max_length = max(
                        max_length,
                        len(str(cell.value))
                    )
            except:
                pass
        ws.column_dimensions[col_letter].width = max_length + 3
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return StreamingResponse(
        buffer,
        media_type=
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition":
            "attachment; filename=sales_report.xlsx"
        }
    )

def format_pk(value):
    if value is None:
        return "0.00"
    return "{:,.2f}".format(float(value))

@router.get("/pdf/{invoice_id}")
async def generate_invoice_pdf(
    invoice_id: int,
    db: Session = Depends(get_db),
    fbr=Depends(get_fbr_client_secure),
):
    client_id = fbr.client_id
    invoice = db.query(Invoice).filter(
        Invoice.id == invoice_id,
        Invoice.client_id == client_id
    ).first()
    if not invoice:
        raise HTTPException(status_code=404, detail="Invoice not found")
    items = db.query(InvoiceItem).filter(
        InvoiceItem.invoice_id == invoice.id
    ).all()
    item_rows = []
    subtotal_excl_st = 0
    retail_price = 0
    total_sales_tax = 0
    total_discount = 0
    total_further_tax = 0
    st_wh = 0
    total_extra_tax = 0
    total_fed = 0
    sale_type = ""
    for item in items:
        subtotal_excl_st += float(item.valueSalesExcludingST)
        retail_price += float(item.fixedNotifiedValueOrRetailPrice)
        total_sales_tax += float(item.salesTaxApplicable)
        total_discount += float(item.discount or 0)
        total_further_tax += float(item.furtherTax or 0)
        st_wh += float(item.salesTaxWithheldAtSource)
        total_extra_tax += float(item.extraTax or 0)
        total_fed += float(item.fedPayable or 0)
        sale_type += str(item.saleType)
        item_rows.append({
            "hs_code": item.hsCode,
            "description": item.productDescription,
            "sale_type": item.saleType,
            "quantity": float(item.quantity),
            "sro_sched": item.sroScheduleNo,
            "sro_item_sr": item.sroItemSerialNo,
            "sale_type": item.saleType,
            "uom": item.uom,
            "rate": item.rate,
            "value_excl": float(item.valueSalesExcludingST),
            "sales_tax": float(item.salesTaxApplicable),
            "total": float(item.totalValues)
        })
    grand_total = (
        subtotal_excl_st
        + total_sales_tax
        + total_further_tax
        + total_extra_tax
        + total_fed
        - total_discount
        - st_wh
    )
    qr_img = qrcode.make(invoice.fbrInvoiceNo)
    qr_buffer = BytesIO()
    qr_img.save(qr_buffer, format="PNG")
    qr_buffer.seek(0)
    data = {
        "seller_name": invoice.sellerBusinessName,
        "seller_address": invoice.sellerAddress,
        "seller_ntn": invoice.sellerNTNCNIC,
        "province": invoice.sellerProvince,
        "buyer_name": invoice.buyerBusinessName,
        "buyer_address": invoice.buyerAddress,
        "buyer_ntn": invoice.buyerNTNCNIC,
        "invoice_no": invoice.fbrInvoiceNo,
        "date": invoice.invoiceDate.strftime('%d-%m-%Y'),
        "sale_type": invoice.invoiceType,
        "items": [{
            "hs_code": i["hs_code"],
            "description": i["description"],
            "sale_type": i["sale_type"],
            "quantity": i["quantity"],
            "uom": i["uom"],
            "rate": i["rate"],
            "sro": i["sro_sched"],
            "sro_item": i["sro_item_sr"],
            "value_excl": i["value_excl"],
        } for i in item_rows],
        "sales_tax": total_sales_tax,
        "further_tax": total_further_tax,
        "extra_tax": total_extra_tax,
        "fed": total_fed,
        "discount": total_discount,
        "retail_price": retail_price,
        "st_wh": st_wh,
        "grand_total": grand_total,
        "qr": qr_buffer,
        "logo": str(BASE_DIR / "static/images/fbr_logo.png")
    }
    pdf_bytes = generate_invoice_pdf_rl(data)
    return StreamingResponse(
        BytesIO(pdf_bytes),
        media_type="application/pdf",
        headers={
            "Content-Disposition": f"attachment; filename=invoice_{invoice.fbrInvoiceNo}.pdf"
        }
    )